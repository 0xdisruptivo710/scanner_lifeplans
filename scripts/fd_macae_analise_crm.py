"""
FD Macaé — análise CRM offline a partir de Relatório de atendimentos + Histórico.

Lê:
  - "Relatório de atendimentos.xlsx" (1 linha por atendimento; tem tags/utm/agente)
  - "Histórico de conversas/*.xlsx"   (1 linha por mensagem)

Produz "FD_Macae_Analise_CRM_2026.xlsx" com:
  - Aba "Re-etiquetagem"  : tags sugeridas por contato
  - Aba "Kanban"          : coluna sugerida no Painel Comercial
  - Aba "Sequências"      : sequências de FU sugeridas
  - Aba "Insights Disparos": segmentos para campanha
  - Aba "Resumo"          : distribuição geral

Mapeamento de colunas SOMENTE para as etapas que existem no painel
"Painel Comercial" da WTS Macaé (extraídas em 2026-05-14 via
GET /crm/v1/panel/<id>?IncludeDetails=Steps):
  Recepção - Novo Lead
  Recepção - Atendimento Iniciado
  D1, D2, D3, D4
  Recepção - Avaliação Agendada
  Resgate - Avaliação Realizada
  recepção - Desmarcou Avaliação      (lowercase proposital — bate com WTS)
  Resgate - Recuperação Geral
  Contatos profissionais
"""

import openpyxl, os, re, json
from collections import defaultdict, Counter
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))

# =============================================
# 1. CARREGAR RELATORIO DE ATENDIMENTOS
# =============================================
print("Carregando relatorio de atendimentos...")
wb = openpyxl.load_workbook(os.path.join(BASE, "Relatório de atendimentos.xlsx"), read_only=True)
ws = wb.active

contatos = {}
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    tel = str(row[3]).strip() if row[3] else ""
    nome = str(row[2]).strip() if row[2] else ""
    tags_str = str(row[5]).strip() if row[5] else ""
    agente = str(row[6]).strip() if row[6] else ""
    data_criacao = str(row[10]).strip() if row[10] else ""
    situacao = str(row[17]).strip() if row[17] else ""
    utm_origem = str(row[21]).strip() if row[21] else ""
    utm_campanha = str(row[22]).strip() if row[22] else ""

    if not tel:
        continue

    if tel not in contatos:
        contatos[tel] = {
            "nome": nome, "tags": set(), "atendimentos": 0,
            "situacoes": [], "agentes": set(), "utm_origens": set(),
            "utm_campanhas": set(), "datas": [], "msgs": []
        }
    c = contatos[tel]
    c["atendimentos"] += 1
    if tags_str:
        for t in tags_str.split(","):
            c["tags"].add(t.strip())
    if agente:
        c["agentes"].add(agente)
    c["situacoes"].append(situacao)
    if utm_origem:
        c["utm_origens"].add(utm_origem)
    if utm_campanha:
        c["utm_campanhas"].add(utm_campanha)
    if data_criacao:
        c["datas"].append(data_criacao)
wb.close()
print(f"  {len(contatos)} contatos do relatorio")

# =============================================
# 2. CARREGAR HISTORICO DE CONVERSAS
# =============================================
print("Carregando historico de conversas...")
hist_dir = os.path.join(BASE, "Histórico de conversas")
msg_count = 0
for fname in sorted(os.listdir(hist_dir)):
    fpath = os.path.join(hist_dir, fname)
    wb2 = openpyxl.load_workbook(fpath, read_only=True)
    ws2 = wb2.active
    for j, row in enumerate(ws2.iter_rows(values_only=True)):
        if j == 0:
            continue
        tel = str(row[4]).strip() if row[4] else ""
        nome_msg = str(row[3]).strip() if row[3] else ""
        msg_data = str(row[8]).strip() if row[8] else ""
        quem = str(row[9]).strip() if row[9] else ""
        conteudo = str(row[10]).strip() if row[10] else ""

        if not tel:
            continue

        if tel not in contatos:
            contatos[tel] = {
                "nome": nome_msg,
                "tags": set(), "atendimentos": 0,
                "situacoes": [], "agentes": set(), "utm_origens": set(),
                "utm_campanhas": set(), "datas": [], "msgs": []
            }

        contatos[tel]["msgs"].append({
            "data": msg_data,
            "quem": quem,
            "conteudo": conteudo[:300]
        })
        msg_count += 1
    wb2.close()
    print(f"  {fname}: OK")

print(f"  {msg_count} mensagens carregadas")
print(f"  {len(contatos)} contatos totais")

# =============================================
# 3. ANALISE POR CONTATO - ESTADO ATUAL
# =============================================
print("Analisando estado atual de cada contato...")

KW_AGENDOU = ["agend", "marcado", "confirmad", "te espero", "horario", "horário", "dia e hora"]
KW_COMPROU = ["fechou", "comprou", "pagamento", "pix", "cartao", "cartão", "parcela", "boleto", "aprovado", "contrato"]
KW_DESMARCOU = ["desmarc", "cancel", "nao vai poder", "não vai poder", "remarc", "adiar"]
KW_FALTOU = ["faltou", "não compareceu", "nao compareceu", "no show"]
KW_ORCAMENTO = ["orçamento", "orcamento", "valor", "quanto custa", "preço", "preco", "investimento", "tabela"]
KW_INTERESSE = ["interesse", "quero saber", "informaç", "gostaria", "como funciona", "tem disponib"]
KW_RETORNO = ["retorn", "voltar", "nova sessão", "próxima sessão", "manutençã", "retoqu"]

proc_map = {
    "botox": "Botox", "toxina": "Botox",
    "preenchimento labial": "Preenchimento Labial", "labios": "Preenchimento Labial",
    "bigode": "Preenchimento Bigode Chinês",
    "malar": "Preenchimento Malar",
    "mandíbula": "Preenchimento Mandíbula", "mandibula": "Preenchimento Mandíbula",
    "mento": "Preenchimento Mento",
    "olheir": "Preenchimento Olheira",
    "gluteo": "Preenchimento Glúteos", "glúteo": "Preenchimento Glúteos",
    "ultraformer": "Ultraformer",
    "papada": "Ultraformer Papada",
    "lavieen": "Lavieen",
    "bioestimulador": "Bioestimulador", "bioestimul": "Bioestimulador",
    "sculptra": "Sculptra",
    "skinbooster": "Skinbooster", "skin booster": "Skinbooster",
    "microagulh": "Microagulhamento",
    "fios pdo": "Fios PDO",
    "vasinh": "Vasinhos",
    "lipo enz": "Lipo Enzimática",
    "peeling": "Peeling",
    "drenagem": "Drenagem",
    "depilação": "Epilação", "epilaç": "Epilação",
    "harmonização": "Harmonização",
    "celulite": "Celulite",
    "estrias": "Estrias",
    "clareamento": "Clareamento",
    "rejuvenescimento": "Rejuvenescimento Íntimo",
    "bioglut": "Bioglúteos", "bioglút": "Bioglúteos",
    "perfect": "Perfect Glúteos",
    "limpeza de pele": "Limpeza de Pele",
    "signature": "Signature Lips",
    "pdrn": "PDRN",
    "emagrecimento facial": "Emagrecimento Facial",
    "glow": "Glow & Lift Experience",
}

tag_proc_map = {
    "BOTOX": "Botox", "ULTRAFORMER PAPADA": "Ultraformer Papada",
    "PREENCHIMENTO GLUTEOS": "Preenchimento Glúteos",
    "VASINHOS": "Vasinhos", "BIOGLÚTEOS": "Bioglúteos",
    "LAVIEEN": "Lavieen", "PERFECT GLÚTEOS": "Perfect Glúteos",
    "SIGNATURE LIPS": "Signature Lips", "ESTRIAS": "Estrias",
    "PREENCHIMENTO LABIAL": "Preenchimento Labial",
    "BIOESTIMULADOR": "Bioestimulador", "BIGODE CHINES": "Preenchimento Bigode Chinês",
    "LIPO ENZIMÁTICA": "Lipo Enzimática", "SKINBOOSTER": "Skinbooster",
    "MICROAGULHAMENTO": "Microagulhamento", "CELULITE": "Celulite",
    "HARMONIZAÇÃO": "Harmonização", "SCULPTRA": "Sculptra",
    "ULTRAFORMER FULL FACE": "Ultraformer Full Face",
    "ULTRAFORMER ABDÔMEN": "Ultraformer Abdômen",
    "ULTRAFORMER CORPORAL": "Ultraformer Corporal",
    "PREENCHIMENTO BIGODE CHINÊS": "Preenchimento Bigode Chinês",
    "PREENCHIMENTO MALAR": "Preenchimento Malar",
    "PREENCHIMENTO DE MANDÍBULA": "Preenchimento Mandíbula",
    "PREENCHIMENTO OLHEIRA": "Preenchimento Olheira",
    "PREENCHIMENTO MARIONETE": "Preenchimento Marionete",
    "PREENCHIMENTO CORPORAL": "Preenchimento Corporal",
    "FIOS PDO": "Fios PDO", "PDRN": "PDRN",
    "DEPILAÇÃO": "Epilação", "DRENAGEM": "Drenagem",
    "CLAREAMENTO": "Clareamento", "LIMPEZA DE PELE": "Limpeza de Pele",
    "PEELING": "Peeling", "PEELING RETINÓICO": "Peeling Retinóico",
    "SkinBooster": "Skinbooster", "Bioestimulador": "Bioestimulador",
    "MESOTERAPIA CAPILAR": "Mesoterapia Capilar",
    "EMAGRECIMENTO FACIAL": "Emagrecimento Facial",
    "GLOW & LIFT EXPERIENCE": "Glow & Lift Experience",
    "REJUVENESCIMENTO INTIMO": "Rejuvenescimento Íntimo",
    "PÁLPEBRAS": "Pálpebras",
}

# Mapa estado -> coluna do painel.
# IMPORTANTE: estes targets têm que existir literalmente em
# wts_panel_mapping_macae. As únicas colunas válidas do "Painel Comercial"
# são: Recepção - Novo Lead, Recepção - Atendimento Iniciado, D1..D4,
# Recepção - Avaliação Agendada, Resgate - Avaliação Realizada,
# "recepção - Desmarcou Avaliação" (lowercase!), Resgate - Recuperação Geral,
# Contatos profissionais. Comprou/Fechou e Retorno/Recompra ficam
# mapeados para "Resgate - Avaliação Realizada" (já passaram pela avaliação).
estado_painel = {
    "Lead Novo": "Recepção - Novo Lead",
    "Demonstrou Interesse": "Recepção - Atendimento Iniciado",
    "Orcamento em Aberto": "Recepção - Atendimento Iniciado",
    "Agendou": "Recepção - Avaliação Agendada",
    "Comprou/Fechou": "Resgate - Avaliação Realizada",
    "Desmarcou": "recepção - Desmarcou Avaliação",
    "Faltou": "recepção - Desmarcou Avaliação",
    "Retorno/Recompra": "Resgate - Avaliação Realizada",
    "Sem Resposta": "Resgate - Recuperação Geral",
    "Venda Perdida": "Resgate - Recuperação Geral",
}

# Mapa procedimento -> sequencia
proc_seq_map = {
    "Botox": "Botox",
    "Microagulhamento": "Microagulhamento",
    "Bioestimulador": "Bioestimulador",
    "Sculptra": "Bioestimulador",
    "Preenchimento Labial": "Preenchimento",
    "Preenchimento Bigode Chinês": "Preenchimento",
    "Preenchimento Malar": "Preenchimento",
    "Preenchimento Mandíbula": "Preenchimento",
    "Preenchimento Mento": "Preenchimento",
    "Preenchimento Olheira": "Preenchimento",
    "Preenchimento Glúteos": "Preenchimento",
    "Preenchimento Marionete": "Preenchimento",
    "Preenchimento Corporal": "Preenchimento",
    "Preenchimento": "Preenchimento",
    "Lavieen": "Lavieen",
    "Ultraformer": "Ultraformer",
    "Ultraformer Papada": "Ultraformer",
    "Ultraformer Full Face": "Ultraformer",
    "Ultraformer Abdômen": "Ultraformer",
    "Ultraformer Corporal": "Ultraformer",
    "Skinbooster": "Skinbooster",
    "Fios PDO": "Fios PDO",
    "Epilação": "Epilação",
    "Vasinhos": "Vasinhos",
}

def classify_contact(c):
    msgs = sorted(c["msgs"], key=lambda m: m["data"]) if c["msgs"] else []
    tags_atuais = c["tags"]

    recent = msgs[-20:] if len(msgs) > 20 else msgs

    ultima_data = ""
    if msgs:
        ultima_data = msgs[-1]["data"][:10] if msgs[-1]["data"] else ""
    elif c["datas"]:
        ultima_data = c["datas"][-1][:10] if c["datas"] else ""

    estado = "Lead Novo"
    confianca = "media"

    for msg in reversed(recent):
        txt = msg["conteudo"].lower()
        if any(k in txt for k in KW_COMPROU):
            estado = "Comprou/Fechou"
            confianca = "alta"
            break
        elif any(k in txt for k in KW_AGENDOU):
            estado = "Agendou"
            confianca = "alta"
            break
        elif any(k in txt for k in KW_RETORNO):
            estado = "Retorno/Recompra"
            confianca = "alta"
            break
        elif any(k in txt for k in KW_DESMARCOU):
            estado = "Desmarcou"
            confianca = "alta"
            break
        elif any(k in txt for k in KW_FALTOU):
            estado = "Faltou"
            confianca = "media"
            break
        elif any(k in txt for k in KW_ORCAMENTO):
            estado = "Orcamento em Aberto"
            confianca = "media"
            break
        elif any(k in txt for k in KW_INTERESSE):
            estado = "Demonstrou Interesse"
            confianca = "media"
            break

    if estado == "Lead Novo" and tags_atuais:
        if "VENDA" in tags_atuais:
            estado = "Comprou/Fechou"
            confianca = "media"
        elif "AGENDADO" in tags_atuais:
            estado = "Agendou"
            confianca = "media"
        elif "VENDA PERDIDA" in tags_atuais:
            estado = "Venda Perdida"
            confianca = "media"

    if estado == "Lead Novo" and len(msgs) > 0:
        clinica_msgs = [m for m in recent if "face doctor" in m["quem"].lower()]
        contato_msgs = [m for m in recent if "face doctor" not in m["quem"].lower() and "para: face" in m["quem"].lower()]
        if len(clinica_msgs) > 0 and len(contato_msgs) == 0:
            estado = "Sem Resposta"
            confianca = "media"

    # Procedimentos
    procedimentos = set()
    for msg in msgs:
        txt = msg["conteudo"].lower()
        for kw, proc in proc_map.items():
            if kw in txt:
                procedimentos.add(proc)

    for t in tags_atuais:
        if t in tag_proc_map:
            procedimentos.add(tag_proc_map[t])

    # Sugestao de painel — fallback "Resgate - Recuperação Geral" (também válido)
    painel_sugerido = estado_painel.get(estado, "Resgate - Recuperação Geral")

    # Sugestao de sequencias
    seqs_sugeridas = set()
    if estado in ("Orcamento em Aberto", "Demonstrou Interesse", "Sem Resposta", "Desmarcou", "Venda Perdida"):
        for proc in procedimentos:
            if proc in proc_seq_map:
                seqs_sugeridas.add(proc_seq_map[proc])
        if not seqs_sugeridas:
            seqs_sugeridas.add("Recuperação Geral")

    # Etiquetas sugeridas (indicacao clinica)
    etiquetas_sugeridas = set()
    for proc in procedimentos:
        etiquetas_sugeridas.add(proc)
    # Adicionar etiqueta de status
    status_tags = {
        "Agendou": "AGENDADO",
        "Comprou/Fechou": "Venda",
        "Desmarcou": "Voltar p agendar",
        "Orcamento em Aberto": "orçamento",
        "Sem Resposta": "Sem resposta após tentativas",
        "Retorno/Recompra": "RETORNO",
        "Venda Perdida": "Venda perdida",
    }
    if estado in status_tags:
        etiquetas_sugeridas.add(status_tags[estado])

    return {
        "estado": estado,
        "confianca": confianca,
        "procedimentos": procedimentos,
        "painel_sugerido": painel_sugerido,
        "seqs_sugeridas": seqs_sugeridas,
        "etiquetas_sugeridas": etiquetas_sugeridas,
        "ultima_data": ultima_data,
        "total_msgs": len(msgs),
    }

results = []
for tel, c in contatos.items():
    analysis = classify_contact(c)
    results.append({
        "telefone": tel,
        "nome": c["nome"],
        "tags_atuais": ", ".join(sorted(c["tags"])) if c["tags"] else "",
        "atendimentos": c["atendimentos"],
        "agentes": ", ".join(sorted(c["agentes"])) if c["agentes"] else "",
        "utm_origens": ", ".join(sorted(c["utm_origens"])) if c["utm_origens"] else "",
        "estado_atual": analysis["estado"],
        "confianca": analysis["confianca"],
        "procedimentos": ", ".join(sorted(analysis["procedimentos"])) if analysis["procedimentos"] else "",
        "etiquetas_sugeridas": ", ".join(sorted(analysis["etiquetas_sugeridas"])) if analysis["etiquetas_sugeridas"] else "",
        "painel_sugerido": analysis["painel_sugerido"],
        "seqs_sugeridas": ", ".join(sorted(analysis["seqs_sugeridas"])) if analysis["seqs_sugeridas"] else "",
        "ultima_interacao": analysis["ultima_data"],
        "total_msgs": analysis["total_msgs"],
    })

# =============================================
# 4. GERAR EXCEL
# =============================================
print("\nGerando Excel...")
wb_out = openpyxl.Workbook()

# --- ABA 1: Re-etiquetagem ---
ws1 = wb_out.active
ws1.title = "Re-etiquetagem"
headers1 = ["Telefone", "Nome", "Tags Atuais", "Etiquetas Sugeridas", "Estado Atual", "Confianca", "Procedimentos Detectados", "Ultima Interacao", "Total Msgs"]
ws1.append(headers1)
for r in sorted(results, key=lambda x: x["estado_atual"]):
    ws1.append([
        r["telefone"], r["nome"], r["tags_atuais"], r["etiquetas_sugeridas"],
        r["estado_atual"], r["confianca"], r["procedimentos"],
        r["ultima_interacao"], r["total_msgs"]
    ])

# --- ABA 2: Kanban ---
ws2 = wb_out.create_sheet("Kanban")
headers2 = ["Telefone", "Nome", "Estado Atual", "Coluna Sugerida", "Confianca", "Tags Atuais", "Procedimentos", "Ultima Interacao"]
ws2.append(headers2)
for r in sorted(results, key=lambda x: x["painel_sugerido"]):
    ws2.append([
        r["telefone"], r["nome"], r["estado_atual"], r["painel_sugerido"],
        r["confianca"], r["tags_atuais"], r["procedimentos"], r["ultima_interacao"]
    ])

# --- ABA 3: Sequencias ---
ws3 = wb_out.create_sheet("Sequências")
headers3 = ["Telefone", "Nome", "Estado Atual", "Sequencias Sugeridas", "Procedimentos", "Tags Atuais", "Ultima Interacao"]
ws3.append(headers3)
# Apenas contatos com sequencia sugerida
for r in sorted(results, key=lambda x: x["seqs_sugeridas"]):
    if r["seqs_sugeridas"]:
        ws3.append([
            r["telefone"], r["nome"], r["estado_atual"], r["seqs_sugeridas"],
            r["procedimentos"], r["tags_atuais"], r["ultima_interacao"]
        ])

# --- ABA 4: Insights para Disparos ---
ws4 = wb_out.create_sheet("Insights Disparos")

# Segmentos para campanhas
segmentos = defaultdict(list)
for r in results:
    segmentos[r["estado_atual"]].append(r)

# Contagem de procedimentos para disparos
proc_counter = Counter()
for r in results:
    if r["procedimentos"]:
        for p in r["procedimentos"].split(", "):
            proc_counter[p] += 1

ws4.append(["=== INSIGHTS PARA DISPAROS DE MENSAGENS ===", "", "", ""])
ws4.append([])

# Segmento 1: Orcamento em Aberto
ws4.append(["SEGMENTO 1: ORCAMENTO EM ABERTO", "", "", ""])
ws4.append(["Descricao", "Contatos que receberam orcamento mas nao fecharam", "", ""])
ws4.append(["Total", len(segmentos.get("Orcamento em Aberto", [])), "", ""])
ws4.append(["Acao sugerida", "Disparo de follow-up com condicao especial ou urgencia", "", ""])
ws4.append(["Exemplo de msg", "Oi [nome]! Vi que voce demonstrou interesse em [procedimento]. Temos uma condicao especial essa semana. Posso te contar?", "", ""])
ws4.append([])

# Segmento 2: Sem Resposta
ws4.append(["SEGMENTO 2: SEM RESPOSTA", "", "", ""])
ws4.append(["Descricao", "Contatos que foram abordados mas nao responderam", "", ""])
ws4.append(["Total", len(segmentos.get("Sem Resposta", [])), "", ""])
ws4.append(["Acao sugerida", "Disparo de reativacao com conteudo de valor (antes/depois, depoimentos)", "", ""])
ws4.append(["Exemplo de msg", "Oi [nome]! Passando para compartilhar um resultado incrivel de [procedimento]. Quer saber mais?", "", ""])
ws4.append([])

# Segmento 3: Desmarcou
ws4.append(["SEGMENTO 3: DESMARCOU", "", "", ""])
ws4.append(["Descricao", "Contatos que desmarcaram avaliacao ou procedimento", "", ""])
ws4.append(["Total", len(segmentos.get("Desmarcou", [])), "", ""])
ws4.append(["Acao sugerida", "Disparo de reagendamento com facilidade (link direto, horarios disponiveis)", "", ""])
ws4.append(["Exemplo de msg", "Oi [nome]! Vi que precisou desmarcar. Sem problema! Temos novos horarios disponiveis. Qual dia fica melhor pra voce?", "", ""])
ws4.append([])

# Segmento 4: Venda Perdida
ws4.append(["SEGMENTO 4: VENDA PERDIDA", "", "", ""])
ws4.append(["Descricao", "Contatos classificados como venda perdida", "", ""])
ws4.append(["Total", len(segmentos.get("Venda Perdida", [])), "", ""])
ws4.append(["Acao sugerida", "Campanha de recuperacao com novo beneficio ou condicao", "", ""])
ws4.append([])

# Segmento 5: Retorno/Recompra
ws4.append(["SEGMENTO 5: RETORNO E RECOMPRA", "", "", ""])
ws4.append(["Descricao", "Contatos que ja compraram e podem fazer manutencao ou novo procedimento", "", ""])
ws4.append(["Total", len(segmentos.get("Retorno/Recompra", [])), "", ""])
ws4.append(["Acao sugerida", "Disparo de pos-procedimento e cross-sell de procedimentos complementares", "", ""])
ws4.append([])

# Segmento 6: Faltou
ws4.append(["SEGMENTO 6: FALTOU", "", "", ""])
ws4.append(["Descricao", "Contatos que faltaram a consulta/avaliacao", "", ""])
ws4.append(["Total", len(segmentos.get("Faltou", [])), "", ""])
ws4.append(["Acao sugerida", "Disparo gentil de reagendamento sem tom de cobranca", "", ""])
ws4.append([])

# Top procedimentos para campanha
ws4.append(["=== TOP PROCEDIMENTOS PARA CAMPANHAS ===", "", "", ""])
ws4.append(["Procedimento", "Contatos com Interesse", "Sugestao de Campanha", ""])
for proc, count in proc_counter.most_common(15):
    sugestao = ""
    if "Botox" in proc:
        sugestao = "Campanha Clube Botox / Dia do Botox"
    elif "Ultraformer" in proc:
        sugestao = "Campanha de rejuvenescimento / Flacidez"
    elif "Preenchimento" in proc:
        sugestao = "Campanha harmonizacao facial"
    elif "Glúteos" in proc or "Bioglút" in proc or "Perfect" in proc:
        sugestao = "Campanha corporal gluteos"
    elif "Vasinhos" in proc:
        sugestao = "Campanha vascular"
    elif "Lavieen" in proc:
        sugestao = "Campanha manchas e rejuvenescimento"
    elif "Bioestimulador" in proc or "Sculptra" in proc:
        sugestao = "Campanha colageno e firmeza"
    elif proc == "Skinbooster":
        sugestao = "Campanha hidratacao profunda"
    elif proc == "Epilação":
        sugestao = "Campanha depilacao definitiva"
    else:
        sugestao = f"Campanha segmentada de {proc}"
    ws4.append([proc, count, sugestao, ""])

# UTM analysis
ws4.append([])
ws4.append(["=== ORIGENS DE TRAFEGO ===", "", "", ""])
ws4.append(["Origem", "Contatos", "Recomendacao", ""])
utm_counter = Counter()
for r in results:
    if r["utm_origens"]:
        for o in r["utm_origens"].split(", "):
            utm_counter[o] += 1
    else:
        utm_counter["Organico/Direto"] += 1
for orig, count in utm_counter.most_common():
    ws4.append([orig, count, "", ""])

# --- ABA 5: Resumo ---
ws5 = wb_out.create_sheet("Resumo")
ws5.append(["=== RESUMO GERAL DA ANALISE CRM - FD MACAE ==="])
ws5.append(["Data da analise", datetime.now().strftime("%Y-%m-%d")])
ws5.append(["Total contatos analisados", len(results)])
ws5.append(["Total mensagens processadas", msg_count])
ws5.append([])

ws5.append(["=== DISTRIBUICAO POR ESTADO ==="])
estados = Counter(r["estado_atual"] for r in results)
for e, cnt in estados.most_common():
    ws5.append([e, cnt, f"{cnt/len(results)*100:.1f}%"])

ws5.append([])
ws5.append(["=== CONTATOS POR COLUNA SUGERIDA DO PAINEL ==="])
paineis_count = Counter(r["painel_sugerido"] for r in results)
for p, cnt in paineis_count.most_common():
    ws5.append([p, cnt])

ws5.append([])
ws5.append(["=== SEQUENCIAS SUGERIDAS ==="])
seq_count = Counter()
for r in results:
    if r["seqs_sugeridas"]:
        for s in r["seqs_sugeridas"].split(", "):
            seq_count[s] += 1
for s, cnt in seq_count.most_common():
    ws5.append([s, cnt])

ws5.append([])
ws5.append(["=== CONTATOS SEM TAG (precisam de etiquetagem) ==="])
sem_tag = [r for r in results if not r["tags_atuais"]]
ws5.append(["Total sem tag", len(sem_tag)])

# Salvar
output_path = os.path.join(BASE, "FD_Macae_Analise_CRM_2026.xlsx")
wb_out.save(output_path)
print(f"\nExcel salvo em: {output_path}")

# Estatisticas
print(f"\nESTADOS DETECTADOS:")
for e, cnt in estados.most_common():
    print(f"  {e}: {cnt}")
print(f"\nTOTAL CONTATOS: {len(results)}")
print(f"COM SEQUENCIA SUGERIDA: {len([r for r in results if r['seqs_sugeridas']])}")
print(f"SEM TAG: {len(sem_tag)}")
